"""
Restaurant Leads Scraper
Searches Google Maps for restaurants, scrapes contact info, exports to Excel.

Usage:
    python leads_scraper.py

Dependencies:
    pip install googlemaps requests beautifulsoup4 openpyxl python-dotenv
"""

import os
import re
import time
import googlemaps
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()

API_KEY = os.getenv("GOOGLE_API_KEY")
if not API_KEY:
    raise RuntimeError("GOOGLE_API_KEY not found. Create a .env file with your key.")

gmaps = googlemaps.Client(key=API_KEY)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

SOCIAL_PATTERNS = {
    "instagram": re.compile(r"https?://(www\.)?instagram\.com/[A-Za-z0-9_.]+/?", re.I),
    "facebook": re.compile(r"https?://(www\.)?facebook\.com/[A-Za-z0-9_.%-]+/?", re.I),
}
EMAIL_PATTERN = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.I)
SKIP_EMAIL_DOMAINS = {"sentry.io", "example.com", "wixpress.com", "squarespace.com"}


def search_places(query: str, max_results: int = 60) -> list[dict]:
    """Return up to max_results place summaries for the given query."""
    results = []
    response = gmaps.places(query=query)

    while response and len(results) < max_results:
        for place in response.get("results", []):
            results.append({"place_id": place["place_id"], "name": place["name"]})
            if len(results) >= max_results:
                break

        next_token = response.get("next_page_token")
        if not next_token or len(results) >= max_results:
            break

        time.sleep(2)  # Google requires a short delay before using next_page_token
        response = gmaps.places(query=query, page_token=next_token)

    return results


def get_place_details(place_id: str) -> dict:
    """Fetch phone number and website from Place Details API."""
    fields = ["name", "formatted_phone_number", "website", "international_phone_number"]
    result = gmaps.place(place_id=place_id, fields=fields).get("result", {})
    return {
        "phone": result.get("international_phone_number") or result.get("formatted_phone_number", ""),
        "website": result.get("website", ""),
    }


def scrape_website(url: str) -> dict:
    """Visit a restaurant website and extract email + social links."""
    found = {"email": "", "instagram": "", "facebook": ""}
    if not url:
        return found

    try:
        resp = requests.get(url, headers=HEADERS, timeout=10, allow_redirects=True)
        resp.raise_for_status()
        html = resp.text
        soup = BeautifulSoup(html, "html.parser")

        # Collect all href links
        hrefs = [a.get("href", "") for a in soup.find_all("a", href=True)]
        page_text = " ".join(hrefs) + " " + html

        for platform, pattern in SOCIAL_PATTERNS.items():
            match = pattern.search(page_text)
            if match:
                found[platform] = match.group(0).rstrip("/")

        emails = EMAIL_PATTERN.findall(page_text)
        for email in emails:
            domain = email.split("@")[-1].lower()
            if domain not in SKIP_EMAIL_DOMAINS and not domain.endswith(".png"):
                found["email"] = email
                break

    except Exception:
        pass

    return found


def build_excel(rows: list[dict], filename: str) -> str:
    """Write results to an Excel file and return the path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["Restaurant Name", "Phone", "Email", "Instagram", "Facebook", "Website"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 40

    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row.get("name", ""))
        ws.cell(row=i, column=2, value=row.get("phone", ""))
        ws.cell(row=i, column=3, value=row.get("email", ""))
        ws.cell(row=i, column=4, value=row.get("instagram", ""))
        ws.cell(row=i, column=5, value=row.get("facebook", ""))
        ws.cell(row=i, column=6, value=row.get("website", ""))

    os.makedirs("leads_output", exist_ok=True)
    path = os.path.join("leads_output", filename)
    wb.save(path)
    return path


def run_search(query: str, max_results: int = 60):
    """Main entry point: search, enrich, and export."""
    print(f"\n🔍 Searching: {query}")
    print(f"   Target: {max_results} restaurants\n")

    places = search_places(query, max_results)
    print(f"   Found {len(places)} places. Fetching details...\n")

    rows = []
    for i, place in enumerate(places, start=1):
        print(f"   [{i}/{len(places)}] {place['name']}", end="", flush=True)

        details = get_place_details(place["place_id"])
        web_data = scrape_website(details["website"])

        row = {
            "name": place["name"],
            "phone": details["phone"],
            "website": details["website"],
            **web_data,
        }
        rows.append(row)

        has = [k for k in ("phone", "email", "instagram", "facebook") if row.get(k)]
        print(f"  →  {', '.join(has) if has else 'no contact found'}")

        time.sleep(0.3)

    safe_name = re.sub(r"[^\w\s-]", "", query).strip().replace(" ", "_")
    filename = f"{safe_name}.xlsx"
    path = build_excel(rows, filename)

    found_email = sum(1 for r in rows if r.get("email"))
    found_ig = sum(1 for r in rows if r.get("instagram"))
    found_fb = sum(1 for r in rows if r.get("facebook"))
    found_phone = sum(1 for r in rows if r.get("phone"))

    print(f"\n✅ Done! Saved to: {path}")
    print(f"   Phone: {found_phone} | Email: {found_email} | IG: {found_ig} | FB: {found_fb}")
    return path


if __name__ == "__main__":
    print("=" * 55)
    print("  Restaurant Leads Scraper")
    print("=" * 55)
    print("Examples: 'cafes in Kuala Lumpur', 'restaurants in Tokyo'")
    print("Type 'quit' to exit.\n")

    while True:
        query = input("Search query: ").strip()
        if query.lower() in ("quit", "exit", "q"):
            break
        if not query:
            continue

        try:
            max_r = input("How many results? (default 60, max 300): ").strip()
            max_r = int(max_r) if max_r.isdigit() else 60
            max_r = min(max_r, 300)
        except ValueError:
            max_r = 60

        run_search(query, max_results=max_r)
        print()
