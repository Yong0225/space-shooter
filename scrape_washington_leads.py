"""
Quick lead scraper — Washington DC restaurants
Usage: py scrape_washington_leads.py
Output: leads_output/washington_restaurants.xlsx
"""

import os, re, time, sys
import googlemaps
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY", "")
if not API_KEY:
    sys.exit("ERROR: GOOGLE_API_KEY not found in .env")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}
SOCIAL = {
    "instagram": re.compile(r"https?://(www\.)?instagram\.com/[A-Za-z0-9_.]+/?", re.I),
    "facebook":  re.compile(r"https?://(www\.)?facebook\.com/[A-Za-z0-9_.%-]+/?", re.I),
}
EMAIL_RE   = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.I)
SKIP_EMAIL = {"sentry.io", "example.com", "wixpress.com", "squarespace.com"}
SUBPAGES   = ["/contact", "/about", "/contact-us", "/about-us", "/links"]


def search_places(gmaps, query, max_results=10):
    results, response = [], gmaps.places(query=query)
    while response and len(results) < max_results:
        for p in response.get("results", []):
            results.append({"place_id": p["place_id"], "name": p["name"]})
            if len(results) >= max_results:
                break
        token = response.get("next_page_token")
        if not token or len(results) >= max_results:
            break
        time.sleep(2)
        response = gmaps.places(query=query, page_token=token)
    return results


def get_website(gmaps, place_id):
    r = gmaps.place(place_id=place_id, fields=["website"]).get("result", {})
    return r.get("website", "")


def _extract(html, found):
    soup = BeautifulSoup(html, "html.parser")
    hrefs = [a.get("href", "") for a in soup.find_all("a", href=True)]
    text = " ".join(hrefs) + " " + html
    for platform, pattern in SOCIAL.items():
        if not found[platform]:
            m = pattern.search(text)
            if m:
                found[platform] = m.group(0).rstrip("/")
    if not found["email"]:
        for email in EMAIL_RE.findall(text):
            domain = email.split("@")[-1].lower()
            if domain not in SKIP_EMAIL and not domain.endswith(".png"):
                found["email"] = email
                break


def scrape_website(url):
    found = {"email": "", "instagram": "", "facebook": ""}
    if not url:
        return found
    base = url.rstrip("/")
    try:
        resp = requests.get(base, headers=HEADERS, timeout=10, allow_redirects=True)
        resp.raise_for_status()
        _extract(resp.text, found)
    except Exception:
        pass
    for subpage in SUBPAGES:
        if found["instagram"] and found["facebook"] and found["email"]:
            break
        try:
            resp = requests.get(base + subpage, headers=HEADERS, timeout=8, allow_redirects=True)
            if resp.status_code == 200:
                _extract(resp.text, found)
        except Exception:
            continue
    return found


def save_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["Restaurant Name", "Website", "Instagram", "Facebook", "Email"]
    widths  = [35, 45, 45, 45, 40]
    hf   = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill = hf, hfill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = w
    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row["name"])
        ws.cell(row=i, column=2, value=row["website"])
        ws.cell(row=i, column=3, value=row["instagram"])
        ws.cell(row=i, column=4, value=row["facebook"])
        ws.cell(row=i, column=5, value=row["email"])
    os.makedirs("leads_output", exist_ok=True)
    wb.save(path)


def main():
    query = "restaurants cafes in Wyoming US"
    print(f"Searching: {query}")
    gmaps = googlemaps.Client(key=API_KEY)
    places = search_places(gmaps, query, max_results=100)
    total = len(places)
    print(f"Found {total} places. Fetching details + scraping websites...\n")

    rows = []
    for i, p in enumerate(places, 1):
        name = p["name"]
        website = get_website(gmaps, p["place_id"])
        print(f"[{i}/{total}] {name}")
        print(f"        Website: {website or '—'}")
        scraped = scrape_website(website)
        row = {
            "name":      name,
            "website":   website,
            "instagram": scraped["instagram"],
            "facebook":  scraped["facebook"],
            "email":     scraped["email"],
        }
        rows.append(row)
        print(f"        IG: {scraped['instagram'] or '—'}")
        print(f"        FB: {scraped['facebook'] or '—'}")
        print(f"        Email: {scraped['email'] or '—'}\n")
        time.sleep(0.5)

    out = os.path.abspath("leads_output/wyoming_restaurants_cafes.xlsx")
    save_excel(rows, out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
