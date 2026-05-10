import sys, io, json, urllib.request, urllib.error, os, time, re
from dotenv import load_dotenv

# Fix: force UTF-8 stdout so Chinese characters don't crash on Windows cp1252
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

load_dotenv()
API_KEY = os.getenv("GEMINI_API_KEY", "")
if not API_KEY:
    print("ERROR: GEMINI_API_KEY not found in .env")
    sys.exit(1)

MODELS = [
    "gemini-2.5-flash",
    "gemini-2.5-flash-lite",
]

# Usage: py generate_emails.py <input_xlsx>
# Output file is auto-named: <input_basename>_emails.xlsx in the same folder
if len(sys.argv) < 2:
    print("Usage: py generate_emails.py <path/to/qualified.xlsx>")
    sys.exit(1)

INPUT_FILE = sys.argv[1]
base, ext = os.path.splitext(INPUT_FILE)
OUTPUT_FILE = base + "_emails" + ext


def call_gemini(business_name, pain_point):
    prompt = f"""You are a casual, sharp cold email writer for a content studio called Y-Studio that creates high-end visual content for restaurants and cafes.
Write 2 different cold EMAIL versions for a restaurant/cafe called "{business_name}".

The observed pain point is: {pain_point}

Each email must follow this exact 4-part structure:

1. HOOK (1 sentence) — Always start with "Are you open to..." followed by a desirable positive outcome for the business (e.g. getting more walk-ins, getting more bookings, filling more tables, pulling more local attention). Do NOT ask about their pain or problems here — focus purely on the upside they want. Vary the positive outcome between v1 and v2. Examples: "Are you open to getting more walk-ins this month?" / "Are you open to pulling more local attention to Burger & Bear?"

2. VALUE (2 sentences max — keep tight) — One sentence: genuine compliment + specific pain point with empathy, woven together. One sentence: Competitor-Based FOMO — similar spots nearby are leveling up their poster content and you'd hate for {business_name} to get overlooked. Be observational, not aggressive. Target ~30 words for this section.

3. GIFT (1 sentence) — Casually mention that you put together a poster demo for them as a small gift, purely for their reference. Make it feel zero-pressure — they don't have to do anything with it.

4. OPEN LOOP CTA (1 sentence) — Do NOT ask for a call, a meeting, or permission to send anything. Instead, end with a curiosity gap that makes them want to reply. Example: "I had one specific idea in mind for your [specific area from the pain point] — not sure if it's right for you yet." Stop there. Do not close the loop.

Global rules:
- Format: cold EMAIL body only (no subject line, no sign-off)
- Salutation: start with "Hi {business_name},"
- Tone: casual, real, grounded — like a person texting a business owner, not a corporate pitch
- Total word count: 70–85 words. Count every word before finalising. If over 85, cut sentences. This is a hard limit — do not exceed it.
- No hashtags, no emojis, no buzzwords
- Punctuation: use ONLY commas and periods. No dashes, colons, semicolons, exclamation marks, question marks, parentheses, or any other punctuation symbols whatsoever.
- Never say: "free", "guaranteed", "limited offer", "marketing agency", "social media services", "just", "I wanted to", "reach out", "visual", "visuals"
- Use "poster" or "poster demo" instead of "visual" or "visual demo" — we make food posters, not generic visuals
- Do NOT mention specific food items in the demo or gift sentence — just say "poster demo"
- The 2 versions must feel clearly different — vary the hook angle, the compliment, the FOMO framing, and the open loop ending
- Sound like a real human, not a template

Return ONLY the 2 email bodies separated by "---", no labels, no numbering, no extra text."""

    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}]
    }).encode("utf-8")

    for model in MODELS:
        endpoint = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={API_KEY}"
        for attempt in range(3):
            try:
                req = urllib.request.Request(endpoint, data=payload, headers={"Content-Type": "application/json"})
                with urllib.request.urlopen(req, timeout=60) as resp:
                    result = json.loads(resp.read())
                text = result["candidates"][0]["content"]["parts"][0]["text"].strip()
                parts = text.split("---")
                v1 = parts[0].strip() if len(parts) >= 1 else ""
                v2 = parts[1].strip() if len(parts) >= 2 else ""
                return v1, v2
            except urllib.error.HTTPError as e:
                wait = 5 * (attempt + 1)
                print(f"  [{model}] attempt {attempt+1} failed ({e.code}) — retry in {wait}s")
                time.sleep(wait)
            except Exception as e:
                wait = 5 * (attempt + 1)
                print(f"  [{model}] attempt {attempt+1} failed ({type(e).__name__}) — retry in {wait}s")
                time.sleep(wait)
        print(f"  [{model}] all attempts failed, trying next model...")

    raise RuntimeError("All models exhausted")


def clean_body(text):
    # Strip everything except letters, digits, whitespace, commas, periods, apostrophes
    cleaned = re.sub(r"[^a-zA-Z0-9\s,.\']", "", text)
    cleaned = re.sub(r" {2,}", " ", cleaned)
    return cleaned.strip()


# Resume: load OUTPUT_FILE if exists, else start from INPUT_FILE
import shutil
if not os.path.exists(OUTPUT_FILE):
    shutil.copy2(INPUT_FILE, OUTPUT_FILE)
    print(f"Created output file: {OUTPUT_FILE}")
else:
    print(f"Resuming from existing: {OUTPUT_FILE}")

wb = openpyxl.load_workbook(OUTPUT_FILE)
ws = wb.active
headers = [cell.value for cell in ws[1]]

for col_name in ["Subject", "Email v1", "Email v2"]:
    if col_name not in headers:
        headers.append(col_name)
        ws.cell(row=1, column=len(headers), value=col_name)

subject_col = headers.index("Subject") + 1
v1_col = headers.index("Email v1") + 1
v2_col = headers.index("Email v2") + 1
name_col_key = next((k for k in ["Restaurant Name", "Name", "name"] if k in headers), headers[0])
name_col = headers.index(name_col_key) + 1
pain_col = headers.index("Pain Point") + 1
# Owner column is optional — used for subject line if present
owner_col = headers.index("Owner") + 1 if "Owner" in headers else None

total = ws.max_row - 1
done = 0
skipped = 0

for row_idx in range(2, ws.max_row + 1):
    biz_name = ws.cell(row=row_idx, column=name_col).value
    pain_point = ws.cell(row=row_idx, column=pain_col).value
    if not biz_name or not pain_point:
        continue

    # Skip if Email v1 already filled with a real email (resume checkpoint)
    existing_v1 = ws.cell(row=row_idx, column=v1_col).value or ""
    if existing_v1 and existing_v1.startswith("Hi "):
        skipped += 1
        print(f"[{row_idx-1}/{total}] SKIP (already done): {biz_name}")
        continue

    subject = f"Hi {biz_name}"

    print(f"[{row_idx-1}/{total}] Generating: {biz_name} ...")
    try:
        v1, v2 = call_gemini(biz_name, pain_point)
    except Exception as e:
        print(f"  ERROR: {e} — skipping, will retry on next run")
        continue

    v1 = clean_body(v1)
    v2 = clean_body(v2)

    ws.cell(row=row_idx, column=subject_col, value=subject)
    ws.cell(row=row_idx, column=v1_col, value=v1)
    ws.cell(row=row_idx, column=v2_col, value=v2)

    wb.save(OUTPUT_FILE)  # save immediately after each row
    done += 1

    print(f"  Subject: {subject}")
    print(f"  v1: {v1[:80]}...")
    print(f"  v2: {v2[:80]}...")
    time.sleep(3)  # avoid rate-limiting

print(f"\nDone. {done} generated, {skipped} skipped. Output: {OUTPUT_FILE}")
