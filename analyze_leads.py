"""
Lead Qualifier — ICP Pain Point Analyzer
Primary: Playwright screenshots + Gemini Vision (sees actual posts, poster quality, feed content)
Fallback: Google Search grounding (if page scraping fails)

Usage:
    py analyze_leads.py <input_xlsx> <output_xlsx> [--start N] [--end N]

--start / --end: 1-indexed row numbers counted from the first data row (inclusive)

Dependencies:
    py -m pip install requests openpyxl python-dotenv playwright
    py -m playwright install chromium

Environment (.env):
    GEMINI_API_KEY=<your key>
    FB_EMAIL=<facebook login email>
    FB_PASSWORD=<facebook password>
"""

import os, re, json, sys, time, base64, argparse
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
if not GEMINI_API_KEY:
    print("ERROR: GEMINI_API_KEY not found in .env")
    sys.exit(1)

FB_EMAIL = os.getenv("FB_EMAIL", "")
FB_PASSWORD = os.getenv("FB_PASSWORD", "")
FB_STATE_FILE = ".fb_session.json"

GEMINI_URL = (
    "https://generativelanguage.googleapis.com/v1beta/models/"
    "gemini-2.5-flash:generateContent?key=" + GEMINI_API_KEY
)

RATE_LIMIT_DELAY = 3


# ─── Playwright helpers ───────────────────────────────────────────────────────

def detect_platform(url: str) -> str:
    if "instagram.com" in url:
        return "instagram"
    if "facebook.com" in url or "fb.com" in url:
        return "facebook"
    return "unknown"


def _try_click_first_visible(page, selectors, timeout=3000, delay=600, fallback=None):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.is_visible(timeout=timeout):
                loc.click()
                page.wait_for_timeout(delay)
                return True
        except Exception:
            pass
    if fallback:
        fallback()
    return False


def _fill_first_visible(page, selectors, value, timeout=3000):
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.is_visible(timeout=timeout):
                loc.fill(value)
                return True
        except Exception:
            continue
    return False


def _dismiss_facebook_popups(page):
    return _try_click_first_visible(page, [
        '[data-cookiebanner="accept_button"]',
        'div[aria-label="Allow all cookies"]',
        'button[title="Allow all cookies"]',
        'div[aria-label="Accept all"]',
        'button[aria-label="Accept all"]',
    ], timeout=2000, delay=1200)


def _dismiss_instagram_popups(page):
    if not _try_click_first_visible(page, [
        'div[role="dialog"] button[aria-label="Close"]',
        'div[role="dialog"] svg[aria-label="Close"]',
        'button[aria-label="Close"]',
    ]):
        try:
            page.keyboard.press("Escape")
            page.wait_for_timeout(600)
        except Exception:
            pass


def ensure_facebook_login() -> bool:
    if not FB_EMAIL or not FB_PASSWORD:
        return False
    if os.path.exists(FB_STATE_FILE):
        print("[FB] Reusing saved Facebook session")
        return True

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return False

    print("[FB] Logging into Facebook (first-time setup)...")
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ctx = browser.new_context(
                viewport={"width": 1366, "height": 900},
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                locale="en-US",
            )
            page = ctx.new_page()
            page.goto("https://www.facebook.com/login", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(3000)
            _dismiss_facebook_popups(page)
            page.wait_for_timeout(1500)

            if not _fill_first_visible(page,
                ["#email", 'input[name="email"]', 'input[type="email"]', '[autocomplete="email"]'],
                FB_EMAIL, timeout=5000):
                print("[FB] Could not find email input — scraping without login")
                browser.close()
                return False

            if not _fill_first_visible(page,
                ["#pass", 'input[name="pass"]', 'input[type="password"]'],
                FB_PASSWORD):
                print("[FB] Could not find password input — scraping without login")
                browser.close()
                return False

            if not _try_click_first_visible(page, [
                '[data-testid="royal_login_button"]',
                'button[name="login"]',
                '#loginbutton',
                'button[type="submit"]',
            ]):
                page.keyboard.press("Enter")
            page.wait_for_timeout(6000)

            if "checkpoint" in page.url or "/login" in page.url:
                print("[FB] Login needs verification or failed — scraping without login")
                browser.close()
                return False
            ctx.storage_state(path=FB_STATE_FILE)
            print("[FB] Facebook session saved successfully")
            browser.close()
            return True
    except Exception as e:
        print(f"[FB] Login error: {e}")
        return False


def _to_mbasic_url(url: str) -> str:
    """Convert facebook.com URL to mbasic.facebook.com — no login popup, works without session."""
    return re.sub(r'https?://(www\.)?facebook\.com', 'https://mbasic.facebook.com', url)


def scrape_page(url: str) -> dict:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return {
            "screenshot_b64": None,
            "page_text": "",
            "platform": detect_platform(url),
            "error": "playwright not installed — run: py -m pip install playwright && py -m playwright install chromium",
        }

    platform = detect_platform(url)
    use_mbasic = platform == "facebook" and not os.path.exists(FB_STATE_FILE)
    if use_mbasic:
        url = _to_mbasic_url(url)
        print(f"    [FB] No session — using mbasic: {url[:70]}")

    result = {"screenshot_b64": None, "page_text": "", "platform": platform, "error": None}

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx_kwargs = dict(
            viewport={"width": 1366, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-US",
        )
        if platform == "facebook" and os.path.exists(FB_STATE_FILE):
            ctx_kwargs["storage_state"] = FB_STATE_FILE

        ctx = browser.new_context(**ctx_kwargs)
        page = ctx.new_page()

        try:
            page.goto(url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(3500)

            if platform == "facebook" and use_mbasic:
                # mbasic is text-only — just scroll and wait for content
                page.evaluate("window.scrollBy(0, 600)")
                page.wait_for_timeout(1500)

            elif platform == "facebook" and not use_mbasic:
                _dismiss_facebook_popups(page)
                page.wait_for_timeout(1000)
                page.evaluate("window.scrollBy(0, 480)")
                page.wait_for_timeout(1500)
                try:
                    page.wait_for_selector(
                        '[role="article"], [data-pagelet*="Timeline"], '
                        '[data-visualcompletion="media-vc-image"]',
                        timeout=7000,
                    )
                    page.wait_for_timeout(1000)
                except Exception:
                    pass

            elif platform == "instagram":
                page.wait_for_timeout(1500)
                _dismiss_instagram_popups(page)
                page.wait_for_timeout(800)
                try:
                    page.wait_for_selector("article img, ._aagv img", timeout=6000)
                except Exception:
                    pass

            try:
                result["page_text"] = page.inner_text("body")
            except Exception:
                result["page_text"] = ""

            # Detect FB login-wall: if page is just the login screen, treat as failure
            login_wall = False
            if platform == "facebook":
                pt_lower = result["page_text"].lower()
                login_signals = [
                    "log in to facebook", "log into facebook",
                    "you must log in to continue", "join facebook",
                    "email or phone number", "create new account",
                ]
                if any(s in pt_lower for s in login_signals) and "followers" not in pt_lower and "likes" not in pt_lower:
                    print("    [FB] Login wall detected — will fall back to search")
                    result["error"] = "login_required"
                    login_wall = True

            if not login_wall:
                screenshot_bytes = page.screenshot(type="jpeg", quality=88)
                result["screenshot_b64"] = base64.b64encode(screenshot_bytes).decode()

        except Exception as e:
            result["error"] = str(e)
        finally:
            browser.close()

    return result


# ─── Gemini API ──────────────────────────────────────────────────────────────

ICP_VISION_PROMPT = """
You are a social media analyst evaluating an F&B business's social media page
for a cold email lead generation campaign.

QUALIFYING CRITERIA — lead qualifies if ANY apply:
1. Followers < 50,000
2. Posts roughly once a month or less
3. No new post for more than 1 month
4. Feed dominated by casual phone snapshots (unedited, poor lighting, random composition)
5. Feed is almost entirely Reels with NO food design posters — ONLY counts if followers < 20,000
6. Has food design posters but quality is poor (basic Canva templates, bad fonts/colors, messy layout)
7. Rarely posts food-related content (but has posted some food content before)

DISQUALIFYING CRITERIA — exclude if ANY apply:
- Followers >= 50,000
- Feed is primarily professional photography (studio lighting, sharp focus, consistent visual style)
- Never posted any food content at all
- Feed is all Reels AND followers >= 20,000
- Not an F&B business (e.g. training school, non-food business)
- Last post was more than 6 months ago (account is effectively abandoned)

──────────────────────────────
BUSINESS INFO:
Name: {name}
Platform: {platform}
URL: {url}

PAGE TEXT — extracted live from the page.
Contains: follower/like counts, post timestamps, like/comment numbers.
---
{page_text}
---
──────────────────────────────

A screenshot of the page is attached.

WHAT TO LOOK FOR IN THE SCREENSHOT:
• Are posts phone snapshots (poor lighting, casual framing) or professional photos?
• Are there design posters? Do they look professionally designed or amateurish (Canva template, messy fonts)?
• Is the feed visually consistent or chaotic?
• What type of content is shown (food photos, Reels thumbnails, random lifestyle)?
• Is there a pop-up or login overlay blocking the content? (If yes, note it but still analyze what IS visible)

Evaluate this account using BOTH the screenshot AND the page text.

Return ONLY valid JSON — no markdown, no code fences, no extra text:
{{
  "qualifies": true or false,
  "disqualify_reason": "reason string if false, else null",
  "pain_points_triggered": [list of criterion numbers that apply, e.g. [1, 4]],
  "followers": approximate follower count as integer or null,
  "last_post_approx": "how long ago was the last post, e.g. '3 weeks ago' or null",
  "pain_point": "1-2 sentence personalized pain point in Simplified Chinese if qualifies=true, else null"
}}

The pain_point field MUST:
- Be Simplified Chinese, 1-2 sentences max — short, direct, specific
- Reference what you ACTUALLY SAW in the screenshot (e.g. specific visual observations)
- Include concrete data from page text where available (follower count, last post date, like counts)
- Use NO second-person address (no 您/你/你们). State facts only.
- NOT be a generic template

Good examples (style reference only — write based on actual observations):
- "Feed 以手机随拍为主，构图随意、光线不均，没有任何设计海报。"
- "最近一次发帖是三个月前，粉丝 1,200，帖子互动基本为零。"
- "有发食物海报，但设计明显是套 Canva 模板，字体混乱，整体视觉很廉价。"
- "Feed 全是 Reels，没有任何食物照片或海报，粉丝仅 800 人。"
- "每周都在更新，但图片都是随手拍，打光和构图都很粗糙，没有品牌感。"
"""

ICP_SEARCH_PROMPT = """
You are a social media analyst for a cold email lead generation campaign targeting F&B businesses.

QUALIFYING CRITERIA (lead qualifies if ANY apply):
1. Followers < 50,000
2. Posts roughly once a month or less
3. No new post for more than 1 month
4. Feed dominated by casual phone snapshots (unedited, random composition)
5. Feed is almost entirely Reels with NO food design posters — ONLY if followers < 20,000
6. Has food design posters but quality is poor (basic Canva templates, bad fonts/colors)
7. Rarely posts food-related content (but has posted some food content before)

DISQUALIFYING CRITERIA (exclude if ANY apply):
- Followers >= 50,000
- Feed is primarily professional photography
- Never posted any food content at all
- Feed is all Reels AND followers >= 20,000
- Not an F&B business
- Last post was more than 6 months ago (account is effectively abandoned)

Use Google Search to find information about this business's social media presence.

Restaurant: {name}
Instagram: {ig_url}
Facebook: {fb_url}

Return ONLY valid JSON — no markdown, no code fences:
{{
  "qualifies": true or false,
  "disqualify_reason": "reason string if false, else null",
  "pain_points_triggered": [list of criterion numbers],
  "followers": approximate follower count as integer or null,
  "last_post_approx": "how long ago was the last post or null",
  "pain_point": "1-2 sentence personalized pain point in Simplified Chinese if qualifies=true, else null"
}}

The pain_point must be specific to this account, mention real observed details, and use no second-person address.
"""


def _extract_gemini_text(data: dict, context: str) -> str:
    candidate = data["candidates"][0]
    parts = candidate.get("content", {}).get("parts", [])
    text_parts = [p["text"] for p in parts if "text" in p]
    if not text_parts:
        raise ValueError(f"No text in {context} (finishReason={candidate.get('finishReason')})")
    return text_parts[-1].strip()


def call_gemini_vision(prompt: str, screenshot_b64: str) -> str:
    payload = {
        "contents": [{
            "parts": [
                {"inline_data": {"mime_type": "image/jpeg", "data": screenshot_b64}},
                {"text": prompt},
            ]
        }],
        "generationConfig": {"temperature": 0.4, "maxOutputTokens": 16384},
    }
    r = requests.post(GEMINI_URL, json=payload, timeout=90)
    r.raise_for_status()
    return _extract_gemini_text(r.json(), "vision")


def call_gemini_search(prompt: str) -> str:
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "tools": [{"google_search": {}}],
        "generationConfig": {"temperature": 0.4, "maxOutputTokens": 16384},
    }
    r = requests.post(GEMINI_URL, json=payload, timeout=90)
    r.raise_for_status()
    return _extract_gemini_text(r.json(), "search")


def parse_json_response(text: str) -> dict:
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        text = match.group(0)
    return json.loads(text)


def analyze_lead(name: str, ig_url: str, fb_url: str) -> dict:
    scrape_result = None
    url_used = None

    for url in filter(None, [fb_url, ig_url]):
        scraped = scrape_page(url)
        if scraped.get("error"):
            print(f"    Scrape fail ({url[:55]}): {scraped['error']}")
            continue
        if scraped.get("screenshot_b64"):
            scrape_result = scraped
            url_used = url
            break

    if scrape_result:
        platform = scrape_result["platform"]
        page_text = scrape_result.get("page_text", "")
        print(f"    Vision analysis ({platform}) — {url_used[:55]}")
        prompt = ICP_VISION_PROMPT.format(
            name=name,
            platform=platform,
            url=url_used,
            page_text=page_text[:4000] if page_text else "(page text unavailable)",
        )
        try:
            return parse_json_response(call_gemini_vision(prompt, scrape_result["screenshot_b64"]))
        except Exception as e:
            print(f"    Vision call failed: {e} — falling back to search")

    print("    Search grounding fallback")
    prompt = ICP_SEARCH_PROMPT.format(name=name, ig_url=ig_url or "N/A", fb_url=fb_url or "N/A")

    try:
        return parse_json_response(call_gemini_search(prompt))
    except json.JSONDecodeError:
        try:
            payload = {
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.4, "maxOutputTokens": 4096},
            }
            r = requests.post(GEMINI_URL, json=payload, timeout=90)
            r.raise_for_status()
            return parse_json_response(r.json()["candidates"][0]["content"]["parts"][0]["text"])
        except Exception as e2:
            return {"qualifies": None, "error": f"JSON parse failed after retry: {e2}"}
    except requests.HTTPError as e:
        detail = ""
        try:
            detail = e.response.json().get("error", {}).get("message", "")
        except Exception:
            pass
        return {"qualifies": None, "error": f"HTTP {e.response.status_code}: {detail or str(e)}"}
    except Exception as e:
        return {"qualifies": None, "error": str(e)}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _auto_detect_column(data_rows, patterns, label):
    for row in data_rows[:10]:
        for ci, val in enumerate(row):
            if val and isinstance(val, str) and any(p in val for p in patterns):
                print(f"Auto-detected {label} column at index {ci}")
                return ci
    return None


def safe_str(s) -> str:
    if s is None:
        return ""
    return str(s).encode("cp1252", "replace").decode("cp1252")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Lead qualifier — ICP pain point analyzer")
    parser.add_argument("input", nargs="?", default="leads_output/mount_austin_cafe_restaurant.xlsx")
    parser.add_argument("output", nargs="?", default="leads_output/mount_austin_qualified.xlsx")
    parser.add_argument("--start", type=int, default=None,
                        help="First data row to process (1-indexed, inclusive)")
    parser.add_argument("--end", type=int, default=None,
                        help="Last data row to process (1-indexed, inclusive)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)

    wb_in = openpyxl.load_workbook(args.input)
    rows = list(wb_in.active.iter_rows(values_only=True))

    HEADER_KEYWORDS = {"restaurant name", "name", "instagram", "facebook", "email", "website", "phone"}
    has_real_header = bool({str(v).strip().lower() for v in rows[0] if v} & HEADER_KEYWORDS)

    if has_real_header:
        header = rows[0]
        data_rows = [r for r in rows[1:] if r[0]]
    else:
        n_cols = len(rows[0]) if rows else 0
        std_names = ["Restaurant Name", "Website", "Google Reviews", "Email", "Facebook"]
        header = tuple(std_names[i] if i < len(std_names) else None for i in range(n_cols))
        data_rows = [r for r in rows if r[0]]

    n_total = len(data_rows)
    print(f"Loaded {n_total} data rows from {args.input}")

    if args.start is not None or args.end is not None:
        s = (args.start - 1) if args.start else 0
        e = args.end if args.end else n_total
        data_rows = data_rows[s:e]
        print(f"Row filter: {args.start or 1}–{args.end or n_total} ({len(data_rows)} rows)")

    col_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}

    def get_col(keys, default=None):
        for k in keys:
            if k in col_map:
                return col_map[k]
        return default

    name_col = get_col(["Restaurant Name", "name"], 0)
    ig_col   = get_col(["Instagram", "instagram", "instagram/facebook"])
    fb_col   = get_col(["Facebook", "facebook"])

    if fb_col is None:
        fb_col = _auto_detect_column(data_rows, ["facebook.com", "fb.com"], "Facebook")
    if ig_col is None:
        ig_col = _auto_detect_column(data_rows, ["instagram.com"], "Instagram")

    meaningful    = [(i, h) for i, h in enumerate(header) if h is not None]
    out_col_indices = [i for i, h in meaningful]
    out_header    = [h for i, h in meaningful] + ["Pain Point"]

    os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else ".", exist_ok=True)

    if os.path.exists(args.output):
        wb_out = openpyxl.load_workbook(args.output)
        ws_out = wb_out.active
        existing_rows = list(ws_out.iter_rows(min_row=2, values_only=True))
        already_done = {str(r[0]).strip() for r in existing_rows if r[0]}
        qualified_count = len(existing_rows)
        print(f"Resuming — {qualified_count} already saved in {args.output}")
    else:
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Qualified Leads"
        ws_out.append(out_header)
        for cell in ws_out[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill(fill_type="solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        wb_out.save(args.output)
        already_done = set()
        qualified_count = 0

    progress_file = args.output + ".progress.json"
    if os.path.exists(progress_file):
        try:
            with open(progress_file, "r", encoding="utf-8") as _pf:
                _prev = set(json.load(_pf))
            extras = _prev - already_done
            if extras:
                print(f"  (+{len(extras)} disqualified/errored from previous run — will skip)")
            already_done |= _prev
        except Exception:
            print(f"  (progress file corrupted — ignoring, will reprocess previous non-qualified leads)")

    def _finalize():
        for col_cells in ws_out.columns:
            max_len    = max((len(str(c.value or "")) for c in col_cells), default=10)
            col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
            ws_out.column_dimensions[col_letter].width = min(max_len + 4, 80)
        pain_col_idx = len(out_header)
        for row_cells in ws_out.iter_rows(min_row=2, min_col=pain_col_idx, max_col=pain_col_idx):
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        wb_out.save(args.output)

    ensure_facebook_login()

    errors = []
    total = len(data_rows)
    for idx, row in enumerate(data_rows, 1):
        name   = (row[name_col] if name_col is not None else row[0]) or ""
        ig_url = (row[ig_col]   if ig_col  is not None else "") or ""
        fb_url = (row[fb_col]   if fb_col  is not None else "") or ""
        if ig_col is not None and fb_col is None:
            fb_url = ig_url

        if not ig_url and not fb_url:
            print(f"[{idx}/{total}] SKIP (no social link): {safe_str(name)}")
            continue

        if str(name).strip() in already_done:
            print(f"[{idx}/{total}] SKIP (already done): {safe_str(name)}")
            continue

        print(f"[{idx}/{total}] Analyzing: {safe_str(name)}")
        result = analyze_lead(name, ig_url, fb_url)

        qualifies = result.get("qualifies")
        triggers  = result.get("pain_points_triggered", [])
        followers = result.get("followers")

        if qualifies is True:
            pain_point = result.get("pain_point", "")
            print(f"  QUALIFIED | Followers: {followers} | Triggers: {triggers}")
            print(f"  {safe_str(str(pain_point)[:120])}")
            ws_out.append([row[i] if i < len(row) else None for i in out_col_indices] + [pain_point])
            qualified_count += 1
            wb_out.save(args.output)  # quick save for resume; column widths set once at end
        elif qualifies is False:
            print(f"  DISQUALIFIED | Followers: {followers} | {safe_str(result.get('disqualify_reason', ''))}")
        else:
            err = result.get("error", "unknown")
            print(f"  ERROR: {safe_str(str(err))}")
            errors.append((name, err))
        name_key = str(name).strip()
        if name_key:
            already_done.add(name_key)
            with open(progress_file, "w", encoding="utf-8") as _pf:
                json.dump(list(already_done), _pf, ensure_ascii=False)

        if idx < total:
            time.sleep(RATE_LIMIT_DELAY)

    _finalize()
    print(f"\n{'='*50}")
    print(f"Done. {qualified_count} qualified leads saved to {args.output}")
    if errors:
        print(f"Errors ({len(errors)}):")
        for n, err in errors:
            print(f"  - {safe_str(n)}: {safe_str(str(err))}")


if __name__ == "__main__":
    main()
