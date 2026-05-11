#!/usr/bin/env python3
"""
OTR Leads Scraper
Collects restaurant/cafe leads from Over-the-Rhine, Cincinnati, OH
Output: OTR leads.xlsx  (Name | Website | Email | Instagram | Facebook)

Resume-safe: writes to Excel + otr_progress.json after EVERY lead.
Run: py scrap.py
      py scrap.py --reset   # clear progress and restart
"""

import re, json, time, random, sys, argparse, io
from pathlib import Path

# Force UTF-8 output on Windows so special chars don't crash
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Config ────────────────────────────────────────────────────────────────────
TARGET      = 9999
OUTPUT      = "Dublin leads.xlsx"
PROGRESS    = "dublin_progress.json"
HEADLESS    = False   # keep visible so you can solve CAPTCHAs / intervene

# Queries for Dublin, Ohio
SEARCH_QUERIES = [
    "restaurants Dublin Ohio",
    "cafe Dublin Ohio",
    "coffee shop Dublin Ohio",
    "food Dublin Ohio",
    "brunch Dublin Ohio",
    "breakfast Dublin Ohio",
    "pizza Dublin Ohio",
    "bakery Dublin Ohio",
    "sandwich Dublin Ohio",
    "dessert Dublin Ohio",
    "asian restaurant Dublin Ohio",
    "mediterranean Dublin Ohio",
    "mexican restaurant Dublin Ohio",
    "italian restaurant Dublin Ohio",
    "sushi Dublin Ohio",
    "burger Dublin Ohio",
    "bar and grill Dublin Ohio",
    "brewery Dublin Ohio",
    "seafood Dublin Ohio",
    "vegan Dublin Ohio",
    "restaurant Bridge Street Dublin Ohio",
    "cafe Bridge Street Dublin Ohio",
    "restaurant Historic Dublin Ohio",
    "food Historic Dublin Ohio",
    "restaurant Dublin Ohio 43016",
    "cafe Dublin Ohio 43016",
    "food Dublin Ohio 43016",
    "restaurant Dublin Ohio 43017",
    "cafe Dublin Ohio 43017",
    "food Dublin Ohio 43017",
]

BAD_EMAIL_DOMAINS = {
    'sentry.io', 'example.com', 'wixpress.com', 'squarespace.com',
    'wordpress.com', 'godaddy.com', 'shopify.com', 'amazonaws.com',
    'googleapis.com', 'schema.org', 'wix.com', 'weebly.com',
    'cloudflare.com', 'mailchimp.com', 'constantcontact.com',
}
BAD_EMAIL_PREFIXES = {'noreply', 'no-reply', 'donotreply', 'info@example', 'test@'}

# Names containing these phrases are skipped — places that primarily sell alcohol/cocktails
ALCOHOL_SKIP_KEYWORDS = [
    'cocktail bar', 'cocktail lounge', 'cocktail room',
    'wine bar', 'wine lounge',
    'whiskey bar', 'whiskey lounge', 'bourbon bar',
    'spirits bar', 'spirits lounge',
    'speakeasy',
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def is_alcohol_only(name):
    n = name.lower()
    return any(kw in n for kw in ALCOHOL_SKIP_KEYWORDS)

def sleep(mn=1.2, mx=2.8):
    time.sleep(random.uniform(mn, mx))

def ts():
    return time.strftime("%H:%M:%S")

# ── Progress (atomic write to avoid corruption) ───────────────────────────────
def load_progress():
    p = Path(PROGRESS)
    if p.exists():
        try:
            with open(p, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {"query_idx": 0, "places": [], "done_urls": [], "leads": []}

def save_progress(prog):
    tmp = PROGRESS + ".tmp"
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(prog, f, ensure_ascii=False, indent=2)
    Path(tmp).replace(PROGRESS)

# ── Excel ─────────────────────────────────────────────────────────────────────
HEADERS      = ["Name", "Website", "Email", "Instagram", "Facebook"]
HDR_BG       = "2E4057"
ROW_BG       = ["FFFFFF", "EDF2F7"]
EMAIL_BG     = "C6F6D5"   # green tint if email found

# Persistent seen database — records every lead ever scraped so we never duplicate
# across sessions regardless of whether the Excel files still exist.
SEEN_FILE = "scraped_seen.json"

# Folders used ONCE to bootstrap SEEN_FILE if it doesn't exist yet
_BOOTSTRAP_DIRS = ["已发送的leads", "未处理的leads", "leads_output", "."]

def _read_excels_for_seen(dirs):
    """Helper: scan xlsx files in given dirs, return (names_set, emails_set)."""
    names, emails, scanned = set(), set(), []
    for d in dirs:
        dp = Path(d)
        if not dp.exists():
            continue
        for xlsx in dp.glob("*.xlsx"):
            try:
                wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                name_col = email_col = None
                for i, h in enumerate(headers):
                    if h and str(h).strip().lower() == "name":
                        name_col = i
                    if h and str(h).strip().lower() == "email":
                        email_col = i
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    if name_col is not None and row[name_col]:
                        names.add(str(row[name_col]).lower().strip())
                    if email_col is not None and row[email_col]:
                        e = str(row[email_col]).lower().strip()
                        if "@" in e:
                            emails.add(e)
                wb.close()
                scanned.append(str(xlsx))
            except Exception:
                continue
    return names, emails, scanned

def _save_seen_db(names, emails):
    """Atomically persist the seen database."""
    tmp = SEEN_FILE + ".tmp"
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump({"names": sorted(names), "emails": sorted(emails)}, f,
                  ensure_ascii=False, indent=2)
    Path(tmp).replace(SEEN_FILE)

def load_seen_db():
    """Load the persistent seen database.

    If scraped_seen.json exists, use it directly.
    If not (first run), bootstrap once from existing Excel files then save.
    --reset never wipes this file — it is permanent record of all scraped leads.
    """
    p = Path(SEEN_FILE)
    if p.exists():
        try:
            with open(p, encoding='utf-8') as f:
                data = json.load(f)
            names  = set(data.get("names",  []))
            emails = set(data.get("emails", []))
            print(f"[SeenDB] {len(names)} names | {len(emails)} emails — loaded from {SEEN_FILE}")
            return names, emails
        except Exception:
            pass

    # First-ever run: bootstrap from whatever Excel files exist right now
    print(f"[SeenDB] {SEEN_FILE} not found — bootstrapping from existing lead files (one-time)...")
    names, emails, scanned = _read_excels_for_seen(_BOOTSTRAP_DIRS)
    print(f"[SeenDB] Bootstrapped from {len(scanned)} file(s): {len(names)} names | {len(emails)} emails")
    _save_seen_db(names, emails)
    print(f"[SeenDB] Saved to {SEEN_FILE} — future runs will use this file only")
    return names, emails

def load_existing_names():
    """Read business names already in the current OUTPUT Excel to skip duplicates."""
    p = Path(OUTPUT)
    if not p.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(OUTPUT, read_only=True)
        ws = wb.active
        names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                names.add(str(row[0]).lower().strip())
        wb.close()
        return names
    except Exception:
        return set()

def count_existing_rows():
    """Return current number of data rows (excluding header) in the Excel."""
    p = Path(OUTPUT)
    if not p.exists():
        return 0
    try:
        wb = openpyxl.load_workbook(OUTPUT, read_only=True)
        ws = wb.active
        count = ws.max_row - 1  # subtract header row
        wb.close()
        return max(count, 0)
    except Exception:
        return 0

def init_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OTR Leads"
    ws.append(HEADERS)
    for col, _ in enumerate(HEADERS, 1):
        c = ws.cell(1, col)
        c.font      = Font(bold=True, color="FFFFFF", size=12)
        c.fill      = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    for col, w in zip("ABCDE", [38, 42, 38, 42, 42]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(OUTPUT)
    print(f"[Excel] Created {OUTPUT}")

def append_lead_to_excel(lead):
    wb = openpyxl.load_workbook(OUTPUT)
    ws = wb.active
    ws.append([
        lead.get("name", ""),
        lead.get("website", ""),
        lead.get("email", ""),
        lead.get("instagram", ""),
        lead.get("facebook", ""),
    ])
    r = ws.max_row
    bg = EMAIL_BG if lead.get("email") else ROW_BG[(r - 2) % 2]
    for c in range(1, 6):
        cell = ws.cell(r, c)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
    wb.save(OUTPUT)

# ── Email / social extraction ─────────────────────────────────────────────────
EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')

def clean_emails(raw_list):
    out = []
    seen = set()
    for e in raw_list:
        e = e.lower().strip().rstrip('.')
        if e in seen:
            continue
        seen.add(e)
        domain = e.split('@')[-1]
        if domain in BAD_EMAIL_DOMAINS:
            continue
        if any(e.startswith(p) for p in BAD_EMAIL_PREFIXES):
            continue
        if len(e) < 6 or '.' not in domain:
            continue
        out.append(e)
    return out

def extract_emails_from_html(html):
    # mailto: links first (most reliable)
    mailto = [m.group(1).split('?')[0] for m in re.finditer(r'mailto:([^\s"\'<>]+)', html)]
    # then plain text matches
    plain  = EMAIL_RE.findall(html)
    return clean_emails(mailto + plain)

def extract_socials(html):
    fb_url, ig_url = None, None
    for m in re.finditer(r'href=["\']([^"\']*facebook\.com/(?!sharer|share|dialog|tr\?|plugins)[^"\']+)["\']', html):
        fb_url = m.group(1).split('?')[0].rstrip('/'); break
    for m in re.finditer(r'href=["\']([^"\']*instagram\.com/(?!p/|reel/)[^"\']+)["\']', html):
        ig_url = m.group(1).split('?')[0].rstrip('/'); break
    return fb_url, ig_url

# ── Google Maps scraping ──────────────────────────────────────────────────────
def scroll_maps_panel(page):
    try:
        panel = page.locator('[role="feed"]').first
        for _ in range(14):
            panel.evaluate("el => el.scrollBy(0, 700)")
            sleep(0.7, 1.3)
    except Exception:
        for _ in range(10):
            try:
                page.keyboard.press("End")
                sleep(0.6, 1.0)
            except Exception:
                break

def scrape_maps_query(page, query):
    url = "https://www.google.com/maps/search/" + query.replace(" ", "+")
    print(f"\n[{ts()}] Maps search: {query}")
    try:
        page.goto(url, timeout=30000, wait_until="domcontentloaded")
        sleep(2.5, 4.0)

        # Dismiss any cookie / consent dialog
        for sel in ['button[aria-label*="Reject"]', 'button[aria-label*="Accept"]',
                    'form[action*="consent"] button']:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=2000):
                    btn.click()
                    sleep(0.5, 1.0)
                    break
            except Exception:
                pass

        scroll_maps_panel(page)
        sleep(1.0, 2.0)

        places = []
        # Primary selector used by Google Maps for place result links
        cards = page.locator('a.hfpxzc').all()
        if not cards:
            cards = page.locator('a[href*="/maps/place/"]').all()

        for card in cards:
            try:
                name = (card.get_attribute('aria-label') or '').strip()
                href = card.get_attribute('href') or ''
                if not name or '/maps/place/' not in href:
                    continue
                if href.startswith('//'):
                    href = 'https:' + href
                elif href.startswith('/'):
                    href = 'https://www.google.com' + href
                places.append({'name': name, 'maps_url': href})
            except Exception:
                continue

        print(f"  -> {len(places)} places found")
        return places

    except Exception as e:
        print(f"  [!] Query failed: {e}")
        return []

# ── Place detail → website ────────────────────────────────────────────────────
def get_website(page, maps_url):
    """Visit a Google Maps place page; return the business website URL or None."""
    try:
        page.goto(maps_url, timeout=25000, wait_until="domcontentloaded")
        sleep(1.5, 2.5)

        # Most reliable: the "Website" button has data-item-id="authority"
        link = page.locator('a[data-item-id="authority"]').first
        if link.count():
            href = link.get_attribute('href') or ''
            if href.startswith('http'):
                return href.split('?')[0]

        # Fallback: any external link shown in the info panel
        for a in page.locator('a[href^="http"]').all():
            href = a.get_attribute('href') or ''
            if href.startswith('http') and 'google' not in href and 'maps' not in href:
                return href.split('?')[0]

    except Exception:
        pass
    return None

# ── Website scraping ──────────────────────────────────────────────────────────
def collect_html(page, base_url):
    """Load homepage + up to 2 contact/about sub-pages; return combined HTML."""
    html_parts = []
    extra_urls = []

    try:
        page.goto(base_url, timeout=15000, wait_until="domcontentloaded")
        sleep(1.0, 2.0)
        html_parts.append(page.content())

        for a in page.locator('a[href]').all()[:60]:
            try:
                href = a.get_attribute('href') or ''
                text = (a.inner_text() or '').lower()
                if any(k in href.lower() or k in text
                       for k in ['contact', 'about', 'reach', 'connect', 'location']):
                    if not href.startswith('http'):
                        href = base_url.rstrip('/') + '/' + href.lstrip('/')
                    if href not in extra_urls and href != base_url:
                        extra_urls.append(href)
            except Exception:
                continue

    except Exception:
        pass

    for url in extra_urls[:2]:
        try:
            page.goto(url, timeout=12000, wait_until="domcontentloaded")
            sleep(0.8, 1.5)
            html_parts.append(page.content())
        except Exception:
            continue

    return "\n".join(html_parts)

def scrape_website(page, website_url):
    try:
        html = collect_html(page, website_url)
        emails = extract_emails_from_html(html)
        fb, ig = extract_socials(html)
        return {
            'email':     emails[0] if emails else None,
            'facebook':  fb,
            'instagram': ig,
        }
    except Exception:
        return {'email': None, 'facebook': None, 'instagram': None}

# ── Facebook About scraping ───────────────────────────────────────────────────
def scrape_fb_email(page, fb_url):
    """Try to pull email from a Facebook page's About section (no login)."""
    try:
        about = fb_url.rstrip('/') + '/about'
        page.goto(about, timeout=20000, wait_until="domcontentloaded")
        sleep(2.0, 3.5)
        html = page.content()

        # Skip if hard login wall (no useful content)
        if html.count('@') == 0 and 'email' not in html.lower():
            return None

        emails = extract_emails_from_html(html)
        return emails[0] if emails else None
    except Exception:
        return None

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--reset', action='store_true', help='Clear progress for this extra run and restart')
    args = parser.parse_args()

    if args.reset:
        for f in [PROGRESS, PROGRESS + ".tmp"]:
            Path(f).unlink(missing_ok=True)
        print("[Reset] Progress cleared.")

    # ── Persistent seen database (never wiped by --reset) ────────────────────
    seen_names, seen_emails = load_seen_db()

    # Also load names in the current OUTPUT Excel (handles mid-run resume)
    existing_names = load_existing_names()
    existing_count = count_existing_rows()
    print(f"[Dedup] {existing_count} leads already in {OUTPUT}")

    # Merge all known names/emails into working sets
    existing_names.update(seen_names)
    existing_emails = set(seen_emails)

    prog   = load_progress()
    leads  = prog.get("leads", [])
    done   = set(prog.get("done_urls", []))
    places = prog.get("places", [])
    q_idx  = prog.get("query_idx", 0)

    # Merge in-progress leads from this run (crash-resume safety)
    for lead in leads:
        existing_names.add(lead['name'].lower().strip())
        if lead.get('email'):
            existing_emails.add(lead['email'].lower().strip())

    if not Path(OUTPUT).exists():
        init_excel()
        for lead in leads:
            append_lead_to_excel(lead)

    print(f"[Start] {len(leads)} new leads done this run | {len(places)} places collected | query #{q_idx}")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=HEADLESS,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
        )
        ctx = browser.new_context(
            viewport={"width": 1280, "height": 800},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = ctx.new_page()

        # ── Phase 1: Collect places from Google Maps ─────────────────────────
        print("\n=== Phase 1: Collecting places from Google Maps ===")
        while len(places) < TARGET + 150 and q_idx < len(SEARCH_QUERIES):
            new_places = scrape_maps_query(page, SEARCH_QUERIES[q_idx])
            existing_place_names = {p['name'].lower() for p in places}
            added = 0
            for p in new_places:
                if p['name'].lower() not in existing_place_names:
                    if is_alcohol_only(p['name']):
                        print(f"  [skip alcohol] {p['name']}")
                        continue
                    places.append(p)
                    existing_place_names.add(p['name'].lower())
                    added += 1
            print(f"  +{added} new unique | total {len(places)}")
            q_idx += 1
            prog.update({"query_idx": q_idx, "places": places})
            save_progress(prog)
            sleep(2.0, 4.0)

        print(f"\n[Phase 1 done] {len(places)} unique places collected")
        if len(places) < 50:
            print("[!] Too few places - Google may be blocking. Try again later or solve CAPTCHA manually.")

        # ── Phase 2: Process each place ──────────────────────────────────────
        print("\n=== Phase 2: Scraping websites + emails ===")
        for place in places:
            if len(leads) >= TARGET:
                break

            maps_url = place['maps_url']
            name     = place['name']

            # Skip already processed in this run
            if maps_url in done:
                continue

            # Skip if name already seen (current file or any past scrape)
            if name.lower().strip() in existing_names:
                print(f"  [skip seen name] {name}")
                done.add(maps_url)
                continue

            print(f"\n[{ts()}] [{len(leads)+1}/{TARGET}] {name}")

            website = get_website(page, maps_url)
            print(f"  Website : {website or '-'}")

            lead = {
                'name':      name,
                'website':   website or '',
                'email':     '',
                'instagram': '',
                'facebook':  '',
            }

            if website:
                data = scrape_website(page, website)
                lead['email']     = data.get('email') or ''
                lead['facebook']  = data.get('facebook') or ''
                lead['instagram'] = data.get('instagram') or ''
                print(f"  Email   : {lead['email'] or '-'}")
                print(f"  FB      : {lead['facebook'] or '-'}")
                print(f"  IG      : {lead['instagram'] or '-'}")

                # No email on website -> try Facebook About section
                if not lead['email'] and lead['facebook']:
                    print(f"  -> Checking FB About for email...")
                    fb_email = scrape_fb_email(page, lead['facebook'])
                    if fb_email:
                        lead['email'] = fb_email
                        print(f"  -> FB email found: {fb_email}")

            # Skip if this email was already seen in any past scrape / sent file
            if lead['email'] and lead['email'].lower().strip() in existing_emails:
                print(f"  [skip dup email] {name} — {lead['email']} already recorded")
                done.add(maps_url)
                continue

            # Save immediately — update Excel, progress file, and seen database
            leads.append(lead)
            existing_names.add(name.lower().strip())
            if lead['email']:
                existing_emails.add(lead['email'].lower().strip())
            done.add(maps_url)
            prog.update({"leads": leads, "done_urls": list(done)})
            save_progress(prog)
            append_lead_to_excel(lead)
            _save_seen_db(existing_names, existing_emails)

            sleep(1.5, 3.0)

        browser.close()

    # ── Summary ───────────────────────────────────────────────────────────────
    with_email    = sum(1 for l in leads if l.get('email'))
    with_fb       = sum(1 for l in leads if l.get('facebook'))
    with_ig       = sum(1 for l in leads if l.get('instagram'))
    with_website  = sum(1 for l in leads if l.get('website'))

    print(f"""
=== Done ===
New leads added : {len(leads)}
With website    : {with_website}
With email      : {with_email}
With Facebook   : {with_fb}
With Instagram  : {with_ig}
Total in file   : {existing_count + len(leads)}
Saved to        : {OUTPUT}
""")

if __name__ == "__main__":
    main()
