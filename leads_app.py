"""
Restaurant Leads Finder - Desktop App
Run: python leads_app.py  (or double-click run_app.bat)
Requires: pip install customtkinter googlemaps requests beautifulsoup4 openpyxl python-dotenv
"""

import os, re, time, threading, queue, subprocess, sys
import googlemaps
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv
import customtkinter as ctk
from tkinter import messagebox

load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY", "")

# ── appearance ──────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}
SOCIAL = {
    "instagram": re.compile(r"https?://(www\.)?instagram\.com/[A-Za-z0-9_.]+/?", re.I),
    "facebook":  re.compile(r"https?://(www\.)?facebook\.com/[A-Za-z0-9_.%-]+/?", re.I),
}
EMAIL_RE   = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.I)
SKIP_EMAIL = {"sentry.io", "example.com", "wixpress.com", "squarespace.com"}


# ── scraping helpers ─────────────────────────────────────────
def search_places(gmaps, query, max_results):
    results, response = [], gmaps.places(query=query)
    while response and len(results) < max_results:
        for p in response.get("results", []):
            results.append({"place_id": p["place_id"], "name": p["name"]})
            if len(results) >= max_results:
                break
        token = response.get("next_page_token")
        if not token or len(results) >= max_results:
            break
        time.sleep(2)
        response = gmaps.places(query=query, page_token=token)
    return results

def get_details(gmaps, place_id):
    r = gmaps.place(place_id=place_id,
                    fields=["name","formatted_phone_number",
                            "website","international_phone_number"]
                    ).get("result", {})
    return {
        "phone":   r.get("international_phone_number") or r.get("formatted_phone_number", ""),
        "website": r.get("website", ""),
    }

SUBPAGES_TO_TRY = ["/contact", "/about", "/links", "/contact-us", "/about-us"]


def _extract_from_html(html, found):
    soup = BeautifulSoup(html, "html.parser")
    hrefs = [a.get("href", "") for a in soup.find_all("a", href=True)]
    text = " ".join(hrefs) + " " + html
    for platform, pattern in SOCIAL.items():
        if not found[platform]:
            m = pattern.search(text)
            if m:
                found[platform] = m.group(0).rstrip("/")
    if not found["email"]:
        for email in EMAIL_RE.findall(text):
            domain = email.split("@")[-1].lower()
            if domain not in SKIP_EMAIL and not domain.endswith(".png"):
                found["email"] = email
                break


def scrape_website(url):
    found = {"email": "", "instagram": "", "facebook": ""}
    if not url:
        return found
    base = url.rstrip("/")
    try:
        resp = requests.get(base, headers=HEADERS, timeout=10, allow_redirects=True)
        resp.raise_for_status()
        _extract_from_html(resp.text, found)
    except Exception:
        pass
    if not found["instagram"] or not found["facebook"]:
        for subpage in SUBPAGES_TO_TRY:
            if found["instagram"] and found["facebook"]:
                break
            try:
                resp = requests.get(base + subpage, headers=HEADERS, timeout=8, allow_redirects=True)
                if resp.status_code == 200:
                    _extract_from_html(resp.text, found)
            except Exception:
                continue
    return found

def save_excel(rows, query):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ["Restaurant Name", "Phone", "Email", "Instagram", "Facebook", "Website"]
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(fill_type="solid", fgColor="1F4E79")
    widths = [35, 20, 35, 40, 40, 40]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill = hf, hfill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = w
    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row.get("name",""))
        ws.cell(row=i, column=2, value=row.get("phone",""))
        ws.cell(row=i, column=3, value=row.get("email",""))
        ws.cell(row=i, column=4, value=row.get("instagram",""))
        ws.cell(row=i, column=5, value=row.get("facebook",""))
        ws.cell(row=i, column=6, value=row.get("website",""))
    os.makedirs("leads_output", exist_ok=True)
    safe = re.sub(r"[^\w\s-]", "", query).strip().replace(" ", "_")
    path = os.path.join("leads_output", f"{safe}.xlsx")
    wb.save(path)
    return path


# ── main app ─────────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Restaurant Leads Finder")
        self.geometry("700x620")
        self.resizable(False, False)
        self._queue  = queue.Queue()
        self._running = False
        self._rows   = []
        self._build_ui()
        self.after(100, self._poll_queue)

    # ── UI layout ────────────────────────────────────────────
    def _build_ui(self):
        # Title
        ctk.CTkLabel(self, text="Restaurant Leads Finder",
                     font=ctk.CTkFont(size=22, weight="bold")).pack(pady=(24, 4))
        ctk.CTkLabel(self, text="Search Google Maps · Export to Excel",
                     font=ctk.CTkFont(size=13), text_color="gray").pack(pady=(0, 18))

        # Input card
        card = ctk.CTkFrame(self, corner_radius=12)
        card.pack(padx=30, fill="x")

        ctk.CTkLabel(card, text="Search Query", font=ctk.CTkFont(weight="bold"),
                     anchor="w").pack(padx=20, pady=(16,4), fill="x")
        self.query_entry = ctk.CTkEntry(card, placeholder_text='e.g. "cafes in Tokyo"',
                                        height=40, font=ctk.CTkFont(size=14))
        self.query_entry.pack(padx=20, fill="x")

        row = ctk.CTkFrame(card, fg_color="transparent")
        row.pack(padx=20, pady=12, fill="x")

        ctk.CTkLabel(row, text="Max Results:", font=ctk.CTkFont(weight="bold")).pack(side="left")
        self.max_var = ctk.StringVar(value="60")
        self.max_entry = ctk.CTkEntry(row, textvariable=self.max_var,
                                      width=70, height=36, justify="center")
        self.max_entry.pack(side="left", padx=10)
        ctk.CTkLabel(row, text="(max 300)", text_color="gray").pack(side="left")

        self.start_btn = ctk.CTkButton(card, text="Start Search", height=44,
                                       font=ctk.CTkFont(size=15, weight="bold"),
                                       command=self._on_start)
        self.start_btn.pack(padx=20, pady=(4,16), fill="x")

        # Log
        ctk.CTkLabel(self, text="Progress", font=ctk.CTkFont(weight="bold"),
                     anchor="w").pack(padx=30, pady=(14,4), fill="x")
        self.log_box = ctk.CTkTextbox(self, height=220, font=ctk.CTkFont(family="Courier", size=12))
        self.log_box.pack(padx=30, fill="x")
        self.log_box.configure(state="disabled")

        # Progress bar
        self.progress = ctk.CTkProgressBar(self, height=8)
        self.progress.pack(padx=30, pady=(8,0), fill="x")
        self.progress.set(0)

        # Stats bar
        stats = ctk.CTkFrame(self, corner_radius=10)
        stats.pack(padx=30, pady=12, fill="x")
        stats.columnconfigure([0,1,2,3], weight=1)

        self._stat_vars = {}
        for col, (icon, label) in enumerate([("📞","Phone"),("✉️","Email"),
                                              ("📷","Instagram"),("👍","Facebook")]):
            f = ctk.CTkFrame(stats, fg_color="transparent")
            f.grid(row=0, column=col, padx=10, pady=10)
            ctk.CTkLabel(f, text=icon, font=ctk.CTkFont(size=18)).pack()
            v = ctk.StringVar(value="0")
            self._stat_vars[label.lower()] = v
            ctk.CTkLabel(f, textvariable=v, font=ctk.CTkFont(size=20, weight="bold")).pack()
            ctk.CTkLabel(f, text=label, text_color="gray",
                         font=ctk.CTkFont(size=11)).pack()

        # Open output folder button
        self.open_btn = ctk.CTkButton(self, text="Open Output Folder", height=36,
                                      fg_color="transparent", border_width=1,
                                      text_color=("gray20","gray80"),
                                      command=self._open_folder)
        self.open_btn.pack(padx=30, pady=(0,20), fill="x")

    # ── actions ──────────────────────────────────────────────
    def _on_start(self):
        if self._running:
            self._running = False
            self.start_btn.configure(text="Start Search", fg_color=("#3B8ED0","#1F6AA5"))
            return

        query = self.query_entry.get().strip()
        if not query:
            messagebox.showwarning("Missing Input", "Please enter a search query.")
            return
        if not API_KEY:
            messagebox.showerror("No API Key",
                "GOOGLE_API_KEY not found.\nCreate a .env file with your key.")
            return

        try:
            max_r = min(int(self.max_var.get()), 300)
        except ValueError:
            max_r = 60

        self._rows = []
        self._clear_log()
        self.progress.set(0)
        for k in self._stat_vars:
            self._stat_vars[k].set("0")

        self._running = True
        self.start_btn.configure(text="Stop", fg_color="#D04040")
        threading.Thread(target=self._worker, args=(query, max_r), daemon=True).start()

    def _worker(self, query, max_r):
        self._log(f"Searching: {query}  (target: {max_r})")
        try:
            gm = googlemaps.Client(key=API_KEY)
            places = search_places(gm, query, max_r)
            total = len(places)
            self._log(f"Found {total} places. Fetching details...\n")

            for i, place in enumerate(places, 1):
                if not self._running:
                    self._log("\nStopped by user.")
                    break

                details  = get_details(gm, place["place_id"])
                web_data = scrape_website(details["website"])
                row = {"name": place["name"], **details, **web_data}
                self._rows.append(row)

                tags = [k for k in ("phone","email","instagram","facebook") if row.get(k)]
                tag_str = "  →  " + ", ".join(tags) if tags else "  →  (none)"
                self._log(f"[{i}/{total}]  {place['name']}{tag_str}")
                self._queue.put(("progress", i / total))
                self._queue.put(("stats", self._rows))
                time.sleep(0.3)

            if self._rows:
                path = save_excel(self._rows, query)
                self._log(f"\nSaved → {path}")
                self._queue.put(("done", path))

        except Exception as e:
            self._log(f"\nError: {e}")
        finally:
            self._running = False
            self._queue.put(("idle", None))

    def _poll_queue(self):
        try:
            while True:
                kind, data = self._queue.get_nowait()
                if kind == "log":
                    self._write_log(data)
                elif kind == "progress":
                    self.progress.set(data)
                elif kind == "stats":
                    for k in ("phone","email","instagram","facebook"):
                        n = sum(1 for r in data if r.get(k))
                        self._stat_vars[k].set(str(n))
                elif kind == "done":
                    messagebox.showinfo("Done!", f"Leads saved to:\n{data}")
                elif kind == "idle":
                    self.start_btn.configure(text="Start Search",
                                             fg_color=("#3B8ED0","#1F6AA5"))
        except queue.Empty:
            pass
        self.after(100, self._poll_queue)

    def _log(self, msg):
        self._queue.put(("log", msg + "\n"))

    def _write_log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _open_folder(self):
        path = os.path.abspath("leads_output")
        os.makedirs(path, exist_ok=True)
        subprocess.Popen(f'explorer "{path}"')


if __name__ == "__main__":
    App().mainloop()
